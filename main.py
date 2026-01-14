import tkinter.ttk as ttk
import customtkinter as ctk
from tkinter import messagebox, filedialog
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# 사용자 정의 모듈 import
from data_manager import DataManager
from add_dialog import AddGuestDialog
from settings_dialog import SettingsDialog
from about_dialog import AboutDialog

# --- 초기 설정 ---
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("dark-blue")


class WeddingApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # 1. 윈도우 설정
        self.title("축의금 정산 매니저 - syann97")
        self.geometry("1280x800")
        self.minsize(1100, 700)  # 최소 너비를 조금 늘려서 버튼 짤림 방지

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.font_bold = ("Malgun Gothic", 12, "bold")
        self.font_body = ("Malgun Gothic", 12)

        # 2. DB 연결 및 데이터 로드
        self.db = DataManager()
        self.guest_list = self.db.load_data()
        self.db.load_config()

        # 통계 데이터를 담을 변수 초기화
        self.stats_side = {}
        self.stats_relation = {}

        # # 테스트 완료
        # if not self.guest_list:
        #     self.guest_list = [
        #         {"name": "홍길동", "amount": 100000, "side": "신랑", "relation": "친구", "affiliation": "삼성전자", "meal": 1,
        #          "note": "축하해"},
        #         {"name": "김철수", "amount": 50000, "side": "신부", "relation": "친척", "affiliation": "이모부", "meal": 2,
        #          "note": ""}
        #     ]
        #     self.db.save_data(self.guest_list)

        # UI 배치
        self.create_top_frame()
        self.create_list_frame()
        self.create_bottom_dashboard()

        self.refresh_ui()

    def create_top_frame(self):
        """상단 검색 및 액션 버튼 (레이아웃 개선됨)"""
        self.top_frame = ctk.CTkFrame(self, corner_radius=10, fg_color="transparent")
        self.top_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))

        # 1. 왼쪽: 타이틀
        self.title_label = ctk.CTkLabel(self.top_frame, text="Wedding Manager",  # 공간 절약을 위해 영문 단축
                                        font=("Roboto Medium", 20), text_color=("gray30", "gray70"))
        self.title_label.pack(side="left", padx=(10, 20))

        # 2. 중앙: 검색 그룹 (Frame으로 묶어서 관리)
        search_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        search_frame.pack(side="left", padx=5)

        self.btn_about = ctk.CTkButton(search_frame, text="ⓘ", width=40, height=35,
                                       fg_color="#78909C", hover_color="#546E7A", font=self.font_bold,
                                       command=self.open_about)  # 함수 연결
        self.btn_about.pack(side="left", padx=(0, 5))

        self.search_combo = ctk.CTkComboBox(search_frame, values=["이름", "소속", "관계", "비고"],
                                            width=90, height=35, font=self.font_body)  # 너비 약간 줄임
        self.search_combo.set("이름")
        self.search_combo.pack(side="left", padx=(0, 5))

        self.search_entry = ctk.CTkEntry(search_frame, placeholder_text="검색어...",
                                         width=200, height=35, font=self.font_body)
        self.search_entry.pack(side="left", padx=(0, 5))
        self.search_entry.bind('<Return>', self.search_guest)

        # 검색 버튼
        self.btn_search = ctk.CTkButton(search_frame, text="🔍", width=50, height=35,  # 텍스트 대신 아이콘 느낌으로
                                        fg_color="#546e7a", hover_color="#455a64", font=self.font_bold,
                                        command=self.search_guest)
        self.btn_search.pack(side="left", padx=(0, 5))

        self.btn_reset = ctk.CTkButton(search_frame, text="↺", width=40, height=35,
                                       fg_color="#90A4AE", hover_color="#78909C", font=self.font_bold,
                                       command=self.reset_search)
        self.btn_reset.pack(side="left")

        # 3. 오른쪽: 액션 버튼 그룹 (Frame으로 묶어서 오른쪽 정렬 고정)
        action_frame = ctk.CTkFrame(self.top_frame, fg_color="transparent")
        action_frame.pack(side="right")

        # 설정 (작게)
        self.btn_settings = ctk.CTkButton(action_frame, text="⚙️", width=50, height=35,
                                          fg_color="#607D8B", hover_color="#455a64", font=self.font_bold,
                                          command=self.open_settings)
        self.btn_settings.pack(side="left", padx=5)

        # 엑셀 (중간)
        self.btn_excel = ctk.CTkButton(action_frame, text="📊 엑셀", width=80, height=35,
                                       fg_color="#1E88E5", hover_color="#1565C0", font=self.font_bold,
                                       command=self.export_to_excel)
        self.btn_excel.pack(side="left", padx=5)

        # 추가 (강조)
        self.btn_add = ctk.CTkButton(action_frame, text="+ 추가", width=90, height=35,
                                     fg_color="#2EB086", hover_color="#219F79", font=self.font_bold,
                                     command=self.open_add_dialog)
        self.btn_add.pack(side="left", padx=5)

        # 삭제 (위험 색상)
        self.btn_delete = ctk.CTkButton(action_frame, text="- 삭제", width=90, height=35,
                                        fg_color="#D84315", hover_color="#BF360C", font=self.font_bold,
                                        command=self.delete_guest)
        self.btn_delete.pack(side="left", padx=5)

    def create_list_frame(self):
        """중앙 리스트 UI"""
        self.list_frame = ctk.CTkFrame(self, corner_radius=15)
        self.list_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        self.list_frame.grid_rowconfigure(1, weight=1)
        self.list_frame.grid_columnconfigure(0, weight=1)

        self.lbl_list_title = ctk.CTkLabel(self.list_frame, text="하객 명단 리스트",
                                           font=("Malgun Gothic", 18, "bold"))
        self.lbl_list_title.grid(row=0, column=0, sticky="w", padx=20, pady=(15, 10))

        style = ttk.Style()
        style.theme_use("clam")

        bg_color = "white"
        header_bg = "#343a40"
        header_fg = "white"
        row_selected = "#3B8ED0"

        style.configure("Treeview", background=bg_color, fieldbackground=bg_color, foreground="black",
                        rowheight=35, borderwidth=0, font=("Malgun Gothic", 11))
        style.configure("Treeview.Heading", background=header_bg, foreground=header_fg, relief="flat",
                        font=("Malgun Gothic", 11, "bold"))
        style.map("Treeview", background=[('selected', row_selected)], foreground=[('selected', 'white')])

        columns = ("No", "Name", "Amount", "GuestOf", "Relation", "Affiliation", "Meal", "Note")
        self.tree = ttk.Treeview(self.list_frame, columns=columns, show="headings", style="Treeview")

        headers = [
            ("No", "No", "center", 50),
            ("Name", "이름", "center", 100),
            ("Amount", "금액 (원)", "e", 120),
            ("GuestOf", "대상", "center", 80),
            ("Relation", "관계", "center", 80),
            ("Affiliation", "소속", "w", 150),
            ("Meal", "식권", "center", 60),
            ("Note", "비고", "w", 250)
        ]

        for col, text, anchor, width in headers:
            self.tree.heading(col, text=text)
            self.tree.column(col, anchor=anchor, width=width)

        self.tree.tag_configure("evenrow", background="#f8f9fa")
        self.tree.tag_configure("oddrow", background="white")

        self.scrollbar = ctk.CTkScrollbar(self.list_frame, orientation="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=self.scrollbar.set)

        self.tree.grid(row=1, column=0, sticky="nsew", padx=(20, 5), pady=(0, 20))
        self.scrollbar.grid(row=1, column=1, sticky="ns", padx=(0, 20), pady=(0, 20))

        self.tree.bind("<Double-1>", self.edit_guest)

    def create_bottom_dashboard(self):
        """하단 통계 대시보드 UI"""
        self.bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.bottom_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=20)
        self.bottom_frame.grid_columnconfigure(0, weight=1, uniform="group1")
        self.bottom_frame.grid_columnconfigure(1, weight=1, uniform="group1")

        # 왼쪽: 상세 분류
        self.stats_detail_frame = ctk.CTkFrame(self.bottom_frame, corner_radius=15, border_width=1, border_color="#ddd")
        self.stats_detail_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        stats_header_frame = ctk.CTkFrame(self.stats_detail_frame, fg_color="transparent")
        stats_header_frame.pack(fill="x", padx=15, pady=(10, 5))

        ctk.CTkLabel(stats_header_frame, text="그룹별 상세 통계", font=("Malgun Gothic", 14, "bold")).pack(side="left")

        self.seg_stat_mode = ctk.CTkSegmentedButton(stats_header_frame,
                                                    values=["대상별", "관계별"],
                                                    width=150,
                                                    command=self.update_mini_tree_view)
        self.seg_stat_mode.set("대상별")
        self.seg_stat_mode.pack(side="right")

        mini_cols = ("Group", "Count", "Sum", "Meal")
        self.mini_tree = ttk.Treeview(self.stats_detail_frame, columns=mini_cols, show="headings", height=5)

        self.mini_tree.heading("Group", text="분류");
        self.mini_tree.column("Group", anchor="center", width=80)
        self.mini_tree.heading("Count", text="인원");
        self.mini_tree.column("Count", anchor="center", width=60)
        self.mini_tree.heading("Sum", text="합계 (원)");
        self.mini_tree.column("Sum", anchor="e", width=100)
        self.mini_tree.heading("Meal", text="식권");
        self.mini_tree.column("Meal", anchor="center", width=60)

        self.mini_tree.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        self.mini_tree.tag_configure("evenrow", background="#f8f9fa")
        self.mini_tree.tag_configure("oddrow", background="white")

        # 오른쪽: 요약 카드
        self.summary_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        self.summary_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        self.summary_frame.grid_columnconfigure(0, weight=1)
        self.summary_frame.grid_columnconfigure(1, weight=1)

        self.lbl_total_people = self.create_card(self.summary_frame, 0, 0, "전체 인원", "0 명", "#3B8ED0")
        self.lbl_total_meal = self.create_card(self.summary_frame, 0, 1, "전체 식권", "0 장", "#2CC985")

        total_card = ctk.CTkFrame(self.summary_frame, corner_radius=15, fg_color="#E3F2FD", border_width=0)
        total_card.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)

        ctk.CTkLabel(total_card, text="총 정산 금액", font=("Malgun Gothic", 14, "bold"), text_color="#1565C0").pack(
            side="left", padx=20)
        self.lbl_total_money = ctk.CTkLabel(total_card, text="0 원", font=("Roboto", 28, "bold"), text_color="#0D47A1")
        self.lbl_total_money.pack(side="right", padx=20)

    def create_card(self, parent, row, col, title, initial_value, icon_color):
        card = ctk.CTkFrame(parent, corner_radius=15, fg_color=("white", "#333333"), border_width=2,
                            border_color="#eee")
        card.grid(row=row, column=col, sticky="nsew", padx=5, pady=5)

        bar = ctk.CTkFrame(card, width=5, fg_color=icon_color, corner_radius=0)
        bar.pack(side="left", fill="y", padx=(0, 10))

        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(side="left", fill="both", expand=True, pady=10)

        ctk.CTkLabel(content, text=title, font=("Malgun Gothic", 12), text_color="gray").pack(anchor="w")
        value_label = ctk.CTkLabel(content, text=initial_value,
                                   font=("Roboto Medium", 20), text_color="black",
                                   width=150, anchor="w")
        value_label.pack(anchor="w")
        return value_label

    def refresh_ui(self, data=None):
        """데이터 갱신 및 UI 업데이트"""
        target_list = data if data is not None else self.guest_list

        for item in self.tree.get_children():
            self.tree.delete(item)

        total_count = 0;
        total_money = 0;
        total_meal = 0
        self.stats_side = {}
        self.stats_relation = {}

        for guest in self.guest_list:
            amount = guest.get("amount", 0)
            side = guest.get("side", "미지정")
            relation = guest.get("relation", "미지정")
            meal = guest.get("meal", 0)

            total_count += 1
            total_money += amount
            total_meal += meal

            if side not in self.stats_side: self.stats_side[side] = {"count": 0, "money": 0, "meal": 0}
            self.stats_side[side]["count"] += 1;
            self.stats_side[side]["money"] += amount;
            self.stats_side[side]["meal"] += meal

            if relation not in self.stats_relation: self.stats_relation[relation] = {"count": 0, "money": 0, "meal": 0}
            self.stats_relation[relation]["count"] += 1;
            self.stats_relation[relation]["money"] += amount;
            self.stats_relation[relation]["meal"] += meal

        for i, guest in enumerate(target_list):
            try:
                real_index = self.guest_list.index(guest)
            except ValueError:
                continue

            name = guest.get("name", "")
            amount = guest.get("amount", 0)
            side = guest.get("side", "")
            relation = guest.get("relation", "")
            affiliation = guest.get("affiliation", "")
            meal = guest.get("meal", 0)
            note = guest.get("note", "")

            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=(
                real_index + 1, name, f"{amount:,}", side, relation, affiliation, meal, note
            ), tags=(tag,))

        self.lbl_total_people.configure(text=f"{total_count} 명")
        self.lbl_total_meal.configure(text=f"{total_meal} 장")
        self.lbl_total_money.configure(text=f"{total_money:,} 원")

        self.update_mini_tree_view(None)

    def update_mini_tree_view(self, value):
        for item in self.mini_tree.get_children():
            self.mini_tree.delete(item)

        mode = self.seg_stat_mode.get()
        target_stats = {}

        if mode == "대상별":
            target_stats = self.stats_side
            self.mini_tree.heading("Group", text="구분 (대상)")
        else:
            target_stats = self.stats_relation
            self.mini_tree.heading("Group", text="관계")

        sorted_keys = sorted(target_stats.keys())

        for i, key in enumerate(sorted_keys):
            stat = target_stats[key]
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.mini_tree.insert("", "end", values=(
                key, f"{stat['count']}명", f"{stat['money']:,}", f"{stat['meal']}"
            ), tags=(tag,))

    def open_add_dialog(self):
        dialog = AddGuestDialog(self, self.db)
        self.wait_window(dialog)
        if dialog.guest_data:
            self.guest_list.append(dialog.guest_data)
            self.db.save_data(self.guest_list)
            self.refresh_ui()

    def edit_guest(self, event):
        selected_item = self.tree.selection()
        if not selected_item: return

        values = self.tree.item(selected_item)['values']
        if not values: return

        list_index = int(values[0]) - 1
        if list_index < 0 or list_index >= len(self.guest_list): return

        target_data = self.guest_list[list_index]
        dialog = AddGuestDialog(self, self.db, initial_data=target_data)
        self.wait_window(dialog)

        if dialog.guest_data:
            self.guest_list[list_index] = dialog.guest_data
            self.db.save_data(self.guest_list)
            self.refresh_ui()

    def delete_guest(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("선택 없음", "삭제할 하객을 목록에서 선택해주세요.")
            return

        count = len(selected_items)
        if not messagebox.askyesno("삭제 확인", f"선택한 {count}명의 하객 정보를 정말 삭제하시겠습니까?\n(이 작업은 되돌릴 수 없습니다.)"):
            return

        indices_to_delete = set()
        for item in selected_items:
            values = self.tree.item(item)['values']
            if values:
                real_index = int(values[0]) - 1
                indices_to_delete.add(real_index)

        new_guest_list = []
        for i, guest in enumerate(self.guest_list):
            if i not in indices_to_delete:
                new_guest_list.append(guest)

        self.guest_list = new_guest_list
        self.db.save_data(self.guest_list)
        self.refresh_ui()

    def export_to_excel(self):
        if not self.guest_list:
            messagebox.showwarning("데이터 없음", "저장할 하객 데이터가 없습니다.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="엑셀 파일로 내보내기",
            initialfile="축의금_정산_명부.xlsx"
        )

        if not file_path:
            return

        try:
            total_count = 0;
            total_money = 0;
            total_meal = 0
            stats_side = {};
            stats_relation = {};
            export_list_data = []

            for i, guest in enumerate(self.guest_list, 1):
                name = guest.get("name", "");
                amount = guest.get("amount", 0)
                side = guest.get("side", "미지정");
                relation = guest.get("relation", "미지정")
                affiliation = guest.get("affiliation", "");
                meal = guest.get("meal", 0)
                note = guest.get("note", "")

                total_count += 1;
                total_money += amount;
                total_meal += meal
                if side not in stats_side: stats_side[side] = {"count": 0, "money": 0, "meal": 0}
                stats_side[side]["count"] += 1;
                stats_side[side]["money"] += amount;
                stats_side[side]["meal"] += meal
                if relation not in stats_relation: stats_relation[relation] = {"count": 0, "money": 0, "meal": 0}
                stats_relation[relation]["count"] += 1;
                stats_relation[relation]["money"] += amount;
                stats_relation[relation]["meal"] += meal

                export_list_data.append({
                    "연번": i, "이름": name, "금액": amount, "구분": side,
                    "관계": relation, "소속": affiliation, "식권": meal, "비고": note
                })

            df_list = pd.DataFrame(export_list_data)
            df_total = pd.DataFrame([{"항목": "전체 인원", "값": f"{total_count} 명"}, {"항목": "전체 식권", "값": f"{total_meal} 장"},
                                     {"항목": "총 정산 금액", "값": f"{total_money:,} 원"}])

            side_data = []
            for key in sorted(stats_side.keys()):
                val = stats_side[key]
                side_data.append(
                    {"구분": key, "인원": f"{val['count']}명", "합계(원)": f"{val['money']:,}", "식권": f"{val['meal']}장"})
            df_side = pd.DataFrame(side_data)

            rel_data = []
            for key in sorted(stats_relation.keys()):
                val = stats_relation[key]
                rel_data.append(
                    {"관계": key, "인원": f"{val['count']}명", "합계(원)": f"{val['money']:,}", "식권": f"{val['meal']}장"})
            df_relation = pd.DataFrame(rel_data)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_list.to_excel(writer, sheet_name='하객 명단', index=False)
                row_total = 0
                pd.DataFrame(["[ 전체 요약 ]"]).to_excel(writer, sheet_name='통계_요약', startrow=row_total, index=False,
                                                     header=False)
                df_total.to_excel(writer, sheet_name='통계_요약', startrow=row_total + 1, index=False)
                row_side = row_total + len(df_total) + 4
                pd.DataFrame(["[ 대상별 통계 ]"]).to_excel(writer, sheet_name='통계_요약', startrow=row_side, index=False,
                                                      header=False)
                df_side.to_excel(writer, sheet_name='통계_요약', startrow=row_side + 1, index=False)
                row_rel = row_side + len(df_side) + 4
                pd.DataFrame(["[ 관계별 통계 ]"]).to_excel(writer, sheet_name='통계_요약', startrow=row_rel, index=False,
                                                      header=False)
                df_relation.to_excel(writer, sheet_name='통계_요약', startrow=row_rel + 1, index=False)

                workbook = writer.book
                border_style = Side(border_style="thin", color="000000")
                border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                header_font = Font(bold=True, color="000000")
                center_align = Alignment(horizontal='center', vertical='center')

                def style_sheet(ws, start_row, start_col, end_row, end_col):
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = border
                            cell.alignment = center_align
                            if r == start_row:
                                cell.fill = header_fill
                                cell.font = header_font
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[column].width = (max_length + 4)

                ws1 = writer.sheets['하객 명단']
                style_sheet(ws1, 1, 1, len(df_list) + 1, len(df_list.columns))

                ws2 = writer.sheets['통계_요약']
                style_sheet(ws2, row_total + 2, 1, row_total + 2 + len(df_total), len(df_total.columns))
                style_sheet(ws2, row_side + 2, 1, row_side + 2 + len(df_side), len(df_side.columns))
                style_sheet(ws2, row_rel + 2, 1, row_rel + 2 + len(df_relation), len(df_relation.columns))

                title_font = Font(bold=True, size=11)
                ws2.cell(row=row_total + 1, column=1).font = title_font
                ws2.cell(row=row_side + 1, column=1).font = title_font
                ws2.cell(row=row_rel + 1, column=1).font = title_font

            messagebox.showinfo("저장 완료", f"성공적으로 저장되었습니다!\n{file_path}")

        except PermissionError:
            messagebox.showwarning("저장 실패", "엑셀 파일이 현재 열려있습니다.\n파일을 닫은 후 다시 시도해주세요.")
        except Exception as e:
            messagebox.showerror("저장 실패", f"엑셀 저장 중 오류가 발생했습니다.\n{str(e)}")

    def open_settings(self):
        dialog = SettingsDialog(self, self.db)
        self.wait_window(dialog)
        self.refresh_ui()

    def search_guest(self, event=None):
        category = self.search_combo.get()
        keyword = self.search_entry.get().strip()

        if not keyword:
            self.refresh_ui()
            return

        key_map = {"이름": "name", "소속": "affiliation", "관계": "relation", "비고": "note"}
        target_key = key_map.get(category, "name")

        filtered_list = [
            guest for guest in self.guest_list
            if keyword in str(guest.get(target_key, ""))
        ]

        self.refresh_ui(data=filtered_list)
        print(f"검색 결과: {len(filtered_list)}건")

    def reset_search(self):
        """검색 초기화"""
        self.search_entry.delete(0, "end")
        self.refresh_ui()

    def open_about(self):
        """프로그램 정보 창 열기"""
        dialog = AboutDialog(self)
        self.wait_window(dialog)


if __name__ == "__main__":
    app = WeddingApp()
    app.mainloop()